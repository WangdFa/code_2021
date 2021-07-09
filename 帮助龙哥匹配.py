import pandas as pd
import numpy as np
import re
import xlwt
import xlrd
from xlutils.copy import copy
from openpyxl import load_workbook


# data = pd.read_excel(r"C:\Users\EDZ\Desktop\充电站客户清单.xlsx")
# A = data.A.unique()
# B = data.B.unique()
# a = []
# for i in A:
#     if type(i) == str:
#         for j in B:
#             if re.findall(i+'+',j).__len__() !=0:
#                 a.append(re.findall(i+'+',j))
# AA = pd.DataFrame(data = a).values.T
# result = [i for i in A if i not in AA]
# pd.DataFrame(data = result).to_excel('result.xlsx')
# print()

# data = xlrd.open_workbook(r"C:\Users\EDZ\Desktop\充电站客户清单.xlsx")
# styleBlueBkg = xlwt.easyxf('pattern: pattern solid, fore_colour red;')
# a = data.sheets()[0]
# wb = copy(a)
# ws = wb.get_sheet(0)
# for i in range(a.nrows):
#     if i%2==0:
#         ws.write(i, 0, a.cell(i,0).value, styleBlueBkg)
# wb.save(r"C:\Users\EDZ\Desktop\充电站客户清单1.xlsx")
# writer = pd.ExcelWriter(r'C:\Users\EDZ\Desktop\标注1\标注\高新大工业t.xlsx')
# book = load_workbook(r'C:\Users\EDZ\Desktop\标注1\标注\高新大工业t.xlsx')
# writer.book = book
data = pd.read_excel(r"C:\Users\EDZ\Desktop\标注2\标注\兴平一般工商业.xlsx",sheet_name=0)
pot = pd.read_excel(r"C:\Users\EDZ\Desktop\work\标注1\标注\2021年3月已做交易客户.xlsx",sheet_name=0).户号
result1 = pd.DataFrame(data =np.arange(5).reshape((1,5)),columns=['原户号','现户号','用户名称','年度预估电量','总结算电量'])
result2 = pd.DataFrame(data =np.arange(5).reshape((1,5)),columns=['原户号','现户号','用户名称','年度预估电量','总结算电量'])
for i in range(len(data)):
    if data.loc[i][0] in pot.values:
        result1.loc[i] = [data.loc[i][2],data.loc[i][0],data.loc[i][1],data.loc[i][13],data.loc[i][14]]
    elif data.loc[i][2] in pot.values:
        result1.loc[i] = [data.loc[i][2],data.loc[i][0],data.loc[i][1],data.loc[i][13],data.loc[i][14]]
    else :
        result2.loc[i] = [data.loc[i][2],data.loc[i][0],data.loc[i][1],data.loc[i][13],data.loc[i][14]]
result1.to_excel(r"C:\Users\EDZ\Desktop\1.xlsx", sheet_name='参与交易')
result2.to_excel(r"C:\Users\EDZ\Desktop\1.xlsx", sheet_name='未参与交易')
# writer.save()
print('')


